import os
import django
from string import ascii_uppercase as AU
from itertools import product, chain
from openpyxl import load_workbook

# необходимые настройки
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ni_bsad.settings.development')
django.setup()

from nickel_super_alloys.models import (Alloy, AlloyingElement,
                        OtherProperties, LongTimeStressRupture)


###############################################################################
#Функкции для работы с моделями БД
###############################################################################
def add_alloy(name, **kwargs):
    a = Alloy.objects.get_or_create(name=name)[0]
    a.type_of_alloy = kwargs.get('type_of_alloy', None)
    a.type_of_structure = kwargs.get('type_of_structure', None)
    a.work_temp = kwargs.get('work_temp', None)
    a.save()
    return a

def add_alloying_element(alloy, name, value):
    e = AlloyingElement.objects.get_or_create(alloy=alloy, element=name, value=value)[0]
    e.save()
    return e

def add_long_time_stress_rupture(alloy, temperature, life, stress):
    s = LongTimeStressRupture.objects.get_or_create(alloy=alloy,
        temperature=temperature, life=life, stress=stress)[0]
    s.save()
    return s

def add_other_property(alloy, **kwargs):
    o = OtherProperties.objects.get_or_create(alloy=alloy)[0]
    o.name = kwargs.get('name', None)
    o.description = kwargs.get("description", None)
    o.save()
    return o


# получение порядка буквенного значения
def get_number_of_letter(s):
    """
    get_number_of_letter("F") =  6
    get_number_of_letter("XY") = 649
    """
    res = 0
    for i, j in enumerate(reversed(s)):
        res += 26  ** i * (ord(j) - 64)
    return res

# функция для получения интервала последовательности букв в Excel
def get_interval(a, b):
    """
    функция-генератор
    интервалы a, b - включительно
    [s for s in get_interval("A", "C")] = ['A', 'B', 'C']
    [s for s in get_interval(3, 4)] = ['C', 'D']
    """
    size1 = (i for i in AU)
    size2 = ("".join(i) for i in product(AU, AU))
    if type(a) == str and type(b) == str:
        a, b = list(map(get_number_of_letter, (a,b)))
    all_size = chain(size1, size2)
    for i in all_size:
        g = get_number_of_letter(i)
        if a <= g and g <= b:
            yield i


def populate():
    # открываю базу-данных
    wb = load_workbook('db-elem.xlsx')
    ws = wb['Лист1']
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=87, max_row=5):
        for cell in row:
            cr = str(cell.row)
            if cell.column == "E":
                alloy = add_alloy(ws["B" + cr].value,
                    type_of_alloy=ws["C" + cr].value,
                    type_of_structure=ws["D" + cr].value,
                    work_temp=ws["E" + cr].value)
            elif cell.value:
                name, value = ws[cell.column + '1'].value, cell.value
                if cell.column in get_interval("F", "AC"):
                    add_alloying_element(alloy, name, value)
                elif cell.column in get_interval(30, 33):
                    add_other_property(alloy, name=name, description=value)
                elif cell.column in get_interval("AH", "CI"):
                    life, temperature  = [int(s) for s in name.split("чσ")]
                    add_long_time_stress_rupture(alloy, temperature, life, stress=int(value))

        print(alloy.name, ' added', alloy.slug)
        print(alloy.alloyingelement_set.all().count(), "elements")
        for c in alloy.alloyingelement_set.all():
            print(c)
        print(alloy.longtimestressrupture_set.all().count(), "long time stress rupture properties")
        print("-"*15)

    wb.close()

# Start execution here!
if __name__ == '__main__':
    print("Starting DataBase population script...")
    populate()
    print('Well done')
